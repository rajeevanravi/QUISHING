from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = "users.xlsx"


if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["Username", "Password"])  
    wb.save(EXCEL_FILE)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/snake.html")
def snake():
    return render_template("snake.html")

@app.route("/save", methods=["POST"])
def save():
    username = request.form.get("username")
    password = request.form.get("password")

    # Save to Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([username, password])
    wb.save(EXCEL_FILE)

    return redirect("snake.html")  

if __name__ == "__main__":
    app.run(debug=True)
